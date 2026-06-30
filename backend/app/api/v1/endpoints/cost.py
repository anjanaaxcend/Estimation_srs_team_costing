from __future__ import annotations

import io
from pathlib import Path

from fastapi import APIRouter, File, HTTPException, Response, UploadFile, Depends, Header
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.api.deps import get_current_user
from app.models.user import User, UserHistory
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.schemas.cost import CostEstimationExportRequest
from app.schemas.team import TeamAllocationDocumentResult
from app.schemas.srs import ModelSelection
from app.schemas.axcend import AxcendExcelExportRequest
from app.core.config import settings
from app.services.token_service import check_token_budget, get_effective_api_key, record_token_usage
from app.services.ingestion.utils import normalize_input
from app.services.planning_sync import calculate_cost_totals, is_testing_role, is_deployment_role, is_management_role
from app.services.team_allocation_service import TeamAllocationService
from app.utils.project_name import resolve_project_name

router = APIRouter()


def _styles() -> dict[str, object]:
    thin = Side(border_style="thin", color="CBD5E1")
    return {
        "title_font": Font(name="Calibri", size=18, bold=True, color="000000"),
        "header_font": Font(name="Calibri", size=11, bold=True, color="000000"),
        "bold_font": Font(name="Calibri", size=11, bold=True),
        "body_font": Font(name="Calibri", size=10),
        "title_fill": PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"),
        "header_fill": PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"),
        "alt_fill": PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"),
        "accent_fill": PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid"),
        "border": Border(top=thin, left=thin, right=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
    }


def _team_cost(member) -> float:
    return round(member.count * member.hourly_rate * member.hours_per_member, 2)


@router.post("/analyze-document", response_model=TeamAllocationDocumentResult)
async def analyze_cost_document(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    x_session_id: str | None = Header(None),
) -> TeamAllocationDocumentResult:
    try:
        file_bytes = await file.read()
        normalized = normalize_input(file_bytes, source="file", filename=file.filename)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Failed to read document: {exc}") from exc

    hint_name = Path(file.filename or "uploaded-brief").stem.replace("-", " ").replace("_", " ").title()
    
    provider = "openai"
    byok = get_effective_api_key(db, current_user, provider)
    model_selection = ModelSelection(provider=provider, api_key=byok) if byok else None

    check_token_budget(db, current_user, x_session_id, provider)

    service = TeamAllocationService()
    try:
        result = service.extract_team_allocation_from_document(
            normalized.cleaned_text,
            hint_name=hint_name,
            selected_model=model_selection,
        )
        
        # Only record token usage if the heuristic extraction failed and AI was called
        heuristic_members = service._extract_members_heuristically(normalized.cleaned_text)
        if not heuristic_members:
            estimated_tokens = max(300, len(normalized.cleaned_text) // 4)
            record_token_usage(
                db,
                current_user,
                x_session_id,
                provider,
                settings.openai_srs_model,
                {
                    "total_tokens": estimated_tokens,
                    "prompt_tokens": int(estimated_tokens * 0.6),
                    "completion_tokens": int(estimated_tokens * 0.4)
                },
                "cost",
                hint_name
            )
        return result
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.post("/export-excel")
async def export_cost_excel(
    payload: CostEstimationExportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Response:
    project_name = resolve_project_name(provided_name=payload.project_name, fallback="Project Costing")

    if current_user:
        history = UserHistory(
            user_id=current_user.id,
            action="Downloaded Cost Excel",
            details=f"Exported Cost Excel for: {project_name}"
        )
        db.add(history)
        db.commit()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cost Summary"
    s = _styles()
    totals = calculate_cost_totals(payload)

    sheet.merge_cells("A1:G1")
    for col in range(1, 8):
        cell = sheet.cell(row=1, column=col)
        cell.fill = s["title_fill"]
        cell.border = s["border"]
    sheet["A1"] = f"Project Cost Estimation: {project_name}"
    sheet["A1"].font = s["title_font"]
    sheet["A1"].alignment = s["center"]

    headers = [
        "Role",
        "Count",
        f"Hourly Rate ({payload.currency})",
        "Weekly Hours",
        "Hours / Member",
        f"Cost / Employee ({payload.currency})",
        f"Total ({payload.currency})",
    ]
    for col, text in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col, value=text)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.border = s["border"]
        cell.alignment = s["center"]

    dev_rows = []
    test_rows = []
    deploy_rows = []

    row = 4
    for index, member in enumerate(payload.members):
        role_name = member.role
        if not is_management_role(role_name):
            if is_testing_role(role_name):
                test_rows.append(row)
            elif is_deployment_role(role_name):
                deploy_rows.append(row)
            else:
                dev_rows.append(row)

        values = [
            role_name,
            member.count,
            member.hourly_rate,
            member.weekly_hours,
            member.hours_per_member,
            f"=C{row}*E{row}",
            f"=B{row}*F{row}",
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            cell.font = s["body_font"]
            cell.border = s["border"]
            cell.alignment = s["left"] if col == 1 else s["center"]
            if col in (3, 6, 7):
                cell.number_format = f'"{payload.currency}" #,##0'
            if index % 2 == 0:
                cell.fill = s["alt_fill"]
        row += 1

    member_end_row = row - 1

    salary_total = totals.get("salary_total", 0)
    pm_cost = totals.get("project_management", 0)
    pm_ratio = (pm_cost / salary_total) if salary_total > 0 else 0.10

    effort_subtotal = salary_total + pm_cost + totals.get("misc_total", 0)
    risk_cost = totals.get("risk_contingency", 0)
    risk_ratio = (risk_cost / effort_subtotal) if effort_subtotal > 0 else 0.10

    nego_cost = totals.get("negotiation_buffer", 0)
    nego_ratio = (nego_cost / effort_subtotal) if effort_subtotal > 0 else 0.05

    summary_start = row + 2
    summary_rows = [
        ("Development Total Cost", f"=ROUND(SUM({','.join(f'G{r}' for r in dev_rows)}), 0)" if dev_rows else "=0"),
        ("Testing Total Cost", f"=ROUND(SUM({','.join(f'G{r}' for r in test_rows)}), 0)" if test_rows else "=0"),
        ("Deployment Total Cost", f"=ROUND(SUM({','.join(f'G{r}' for r in deploy_rows)}), 0)" if deploy_rows else "=0"),
        ("Team Salary Total", f"=ROUND(SUM(G4:G{member_end_row}), 0)"),
        (f"Project Management ({round(pm_ratio * 100)}%)", f"=ROUND(G{summary_start+3}*{pm_ratio}, 0)"),
        (f"Risk Contingency ({round(risk_ratio * 100)}%)", f"=ROUND((G{summary_start+3}+G{summary_start+4}+G{summary_start+8})*{risk_ratio}, 0)"),
        (f"Negotiation Buffer ({round(nego_ratio * 100)}%)", f"=ROUND((G{summary_start+3}+G{summary_start+4}+G{summary_start+8})*{nego_ratio}, 0)"),
        ("Company Profit Slabs", int(round(totals.get("profit_total", 0)))),
        ("Miscellaneous", int(round(totals.get("misc_total", 0)))),
        ("Project Total Cost Estimation", f"=ROUND(G{summary_start+3}+G{summary_start+4}+G{summary_start+8}+G{summary_start+5}+G{summary_start+6}, 0)"),
        ("Grand Total", f"=ROUND(G{summary_start+9}+G{summary_start+7}, 0)"),
    ]

    for offset, (label, formula_or_value) in enumerate(summary_rows):
        current_row = summary_start + offset
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        label_cell = sheet.cell(row=current_row, column=1, value=label)
        value_cell = sheet.cell(row=current_row, column=7, value=formula_or_value)
        
        label_cell.font = s["bold_font"]
        value_cell.font = s["bold_font"]
        label_cell.border = s["border"]
        value_cell.border = s["border"]
        label_cell.alignment = s["left"]
        value_cell.alignment = s["center"]
        value_cell.number_format = f'"{payload.currency}" #,##0'
        for col_idx in range(2, 7):
            sheet.cell(row=current_row, column=col_idx).border = s["border"]
            
        if label == "Grand Total":
            label_cell.fill = s["accent_fill"]
            value_cell.fill = s["accent_fill"]
            label_cell.font = Font(name="Calibri", size=11, bold=True, color="000000")
            value_cell.font = Font(name="Calibri", size=11, bold=True, color="000000")

    details_sheet = workbook.create_sheet(title="Additional Costs")
    detail_headers = ["Category", "Label", f"Amount ({payload.currency})"]
    for col, text in enumerate(detail_headers, start=1):
        cell = details_sheet.cell(row=1, column=col, value=text)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.border = s["border"]
        cell.alignment = s["center"]

    detail_row = 2
    additional_items = [
        ("Project Management", [("Management overhead (10%)", totals["project_management"])]),
        ("Risk Contingency", [("Risk Contingency (10%)", totals["risk_contingency"])]),
        ("Negotiation Buffer", [("Negotiation Buffer (5%)", totals["negotiation_buffer"])]),
        ("Profit Slab", [(item.label, item.amount) for item in payload.profit_slabs]),
        ("Miscellaneous", [(item.label, item.amount) for item in payload.miscellaneous_costs if not any(x in item.label.lower() for x in ["risk", "negotiation"])]),
    ]
    for category, items in additional_items:
        for label, amount in items:
            for col, value in enumerate((category, label, amount), start=1):
                cell = details_sheet.cell(row=detail_row, column=col, value=value)
                cell.font = s["body_font"]
                cell.border = s["border"]
                cell.alignment = s["left"] if col < 3 else s["center"]
            detail_row += 1

    for target_sheet in (sheet, details_sheet):
        target_sheet.column_dimensions["A"].width = 30
        target_sheet.column_dimensions["B"].width = 18
        target_sheet.column_dimensions["C"].width = 22
        target_sheet.column_dimensions["D"].width = 18
        target_sheet.column_dimensions["E"].width = 18
        target_sheet.column_dimensions["F"].width = 20
        target_sheet.column_dimensions["G"].width = 18

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"{project_name.replace(' ', '_')}_Cost_Estimation.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# NEW: Export AXCEND format Excel Workbook (3 sheets)
# ---------------------------------------------------------------------------

@router.post("/export-axcend-excel")
async def export_axcend_excel_endpoint(
    payload: AxcendExcelExportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Response:
    """
    Generate and return a beautifully styled 3-sheet AXCEND format Excel workbook
    consisting of:
      1. Module Estimation
      2. Effort Estimation
      3. Cost Estimation
    """
    project_name = resolve_project_name(provided_name=payload.project_name, fallback="Axcend Project")

    if current_user:
        history = UserHistory(
            user_id=current_user.id,
            action="Downloaded Axcend Excel",
            details=f"Exported Axcend format Excel for: {project_name}"
        )
        db.add(history)
        db.commit()

    from app.services.axcend_excel_service import AxcendExcelService
    service = AxcendExcelService()
    
    try:
        output = service.generate_axcend_workbook(payload)
        filename = f"{project_name.replace(' ', '_')}_Axcend_Estimation.xlsx"
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.get("/exchange-rates")
def get_exchange_rates():
    """Fetch live exchange rates from the internet (fallback to standard rates if offline)."""
    import urllib.request
    import json
    try:
        url = "https://open.er-api.com/v6/latest/USD"
        with urllib.request.urlopen(url, timeout=5) as response:
            if response.status == 200:
                data = json.loads(response.read().decode())
                return data.get("rates", {})
    except Exception as e:
        logger.warning("Failed to fetch live exchange rates from internet: %s", e)
    
    # Fallback rates
    return {
        "USD": 1.0,
        "INR": 83.5,
        "EUR": 0.92,
        "GBP": 0.78,
        "AED": 3.67,
        "SGD": 1.35,
        "AUD": 1.50,
        "CAD": 1.37,
        "JPY": 155.0,
    }


