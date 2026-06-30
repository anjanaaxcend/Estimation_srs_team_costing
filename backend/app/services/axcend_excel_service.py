"""AXCEND Excel export service."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.schemas.axcend import AxcendExcelExportRequest, AxcendResourceRow

C_HEADER_BLUE = "FFFFFF"
C_HEADER_WHITE = "000000"
C_SUMMARY_GREY = "F2F4F7"
C_GRAND_TOTAL = "D9D9D9"
C_GRAND_WHITE = "000000"
C_BORDER_COLOUR = "D9D9D9"


def _thin_border() -> Border:
    side = Side(style="thin", color=C_BORDER_COLOUR)
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(colour: str) -> PatternFill | None:
    if colour == "FFFFFF":
        return None
    return PatternFill(start_color=colour, end_color=colour, fill_type="solid")


def _font(bold: bool = False, size: int = 10, colour: str = "000000", name: str = "Calibri") -> Font:
    return Font(name=name, size=size, bold=bold, color=colour)


def _align(h: str = "center", v: str = "center", wrap: bool = True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _apply(cell, font=None, fill=None, align=None, border=None):
    if font:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border


def _style_merged_range(ws, start_row, start_col, end_row, end_col, font=None, fill=None, align=None, border=None):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            _apply(ws.cell(row=row, column=col), font=font, fill=fill, align=align, border=border)


def _set_col_widths(ws, widths: dict[str, float]):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _dev_pct_for_level(level: str, eng_rows: list) -> float:
    """Compute the development fraction for a given level (S3/S2/S1) from the engineering layout rows."""
    dev_rows = [(rn, r) for rn, r in eng_rows if "software development" in (r.activity or "").lower()]
    total_hours = sum(r.input_hours for _, r in dev_rows if r.input_hours)
    if total_hours <= 0:
        return 0.0
    level_hours = sum(r.input_hours for _, r in dev_rows if r.resource_level == level and r.input_hours)
    return round(level_hours / total_hours, 6)


class AxcendExcelService:
    def generate_axcend_workbook(self, req: AxcendExcelExportRequest) -> io.BytesIO:
        wb = Workbook()
        layout = self._plan_sw_effort_layout(req)

        self._build_sheet1_costing(wb, req)
        self._build_sheet2_effort_summary(wb, layout, req)
        self._build_sheet3_sw_effort(wb, req, layout)
        self._build_sheet4_software_efforts(wb, req)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _plan_sw_effort_layout(self, req: AxcendExcelExportRequest) -> dict[str, object]:
        pre_rows = list(req.pre_engineering)

        internal_pct = req.effort_percentages.internal_testing_pct
        client_pct   = req.effort_percentages.client_testing_pct
        deploy_pct   = req.effort_percentages.deployment_pct

        # Read ALL hours from req.engineering (populated by the frontend from display state)
        def _eng_hrs(keyword: str, level: str | None = None) -> float:
            return sum(
                r.input_hours for r in req.engineering
                if keyword in (r.activity or "").lower()
                and (level is None or r.resource_level == level)
            )

        def _eng_exp(keyword: str) -> float:
            for r in req.engineering:
                if keyword in (r.activity or "").lower():
                    return r.experience_years
            return 8.0

        s3_dev_hours         = _eng_hrs("software development", "S3")
        s2_dev_hours         = _eng_hrs("software development", "S2")
        s1_dev_hours         = _eng_hrs("software development", "S1")
        internal_test_hours  = _eng_hrs("internal testing")
        client_test_hours    = _eng_hrs("client testing")
        deployment_hours_val = _eng_hrs("deployment")

        s3_exp = _eng_exp("software development - s3") or _eng_exp("deployment") or 12.0
        s2_exp = _eng_exp("software development - s2") or _eng_exp("internal testing") or 8.0

        eng_rows = [
            AxcendResourceRow(activity="Software Development - S3",            location="India", resource_level="S3", experience_years=s3_exp, input_hours=s3_dev_hours,         section="engineering"),
            AxcendResourceRow(activity="Software Development - S2",            location="India", resource_level="S2", experience_years=s2_exp, input_hours=s2_dev_hours,         section="engineering"),
            AxcendResourceRow(activity="Software Development - S1",            location="India", resource_level="S1", experience_years=2,      input_hours=s1_dev_hours,         section="engineering"),
            AxcendResourceRow(activity=f"Internal Testing ({round(internal_pct * 100)}% of D&D)",  location="India", resource_level="S2", experience_years=s2_exp, input_hours=internal_test_hours,  section="engineering"),
            AxcendResourceRow(activity=f"Client Testing ({round(client_pct   * 100)}% of D&D)",   location="India", resource_level="S2", experience_years=s2_exp, input_hours=client_test_hours,    section="engineering"),
            AxcendResourceRow(activity=f"Deployment ({round(deploy_pct       * 100)}% of D&D)",   location="India", resource_level="S3", experience_years=s3_exp, input_hours=deployment_hours_val, section="engineering"),
        ]

        pre_header_row = 5
        pre_start_row  = pre_header_row + 1
        pre_total_row  = pre_start_row + len(pre_rows)

        dev_header_row = pre_total_row + 2
        dev_start_row  = dev_header_row + 1
        dev_total_row  = dev_start_row + len(eng_rows)
        grand_total_row = dev_total_row + 2

        return {
            "pre_rows":       [(pre_start_row + idx, row) for idx, row in enumerate(pre_rows)],
            "eng_rows":       [(dev_start_row + idx, row) for idx, row in enumerate(eng_rows)],
            "pm_rows":        [],
            "pre_total_row":  pre_total_row,
            "dev_total_row":  dev_total_row,
            "pm_total_row":   dev_total_row,
            "grand_total_row": grand_total_row,
        }


    @staticmethod
    def _sum_cells_formula(sheet_name: str, rows: list[int]) -> str:
        if not rows:
            return "=0"
        refs = ",".join(f"'{sheet_name}'!F{row}" for row in rows)
        return f"=ROUND(SUM({refs}), 0)"

    @staticmethod
    def _activity_label(row: AxcendResourceRow, req: AxcendExcelExportRequest) -> str:
        activity = (row.activity or "").strip()
        lowered = activity.lower()
        pct = req.effort_percentages

        if "software development" in lowered:
            if row.resource_level == "S3":
                return "Software Development - S3"
            if row.resource_level == "S2":
                return "Software Development - S2"
            if row.resource_level == "S1":
                return "Software Development - S1"
        if "internal testing" in lowered and "%" not in activity:
            return f"Internal Testing ({round(pct.internal_testing_pct * 100)}% of D&D)"
        if ("client testing" in lowered or "external review" in lowered) and "%" not in activity:
            return f"Client Testing ({round(pct.client_testing_pct * 100)}% of D&D)"
        if "deployment" in lowered and "%" not in activity:
            return f"Deployment ({round(pct.deployment_pct * 100)}% of D&D)"
        if "project management" in lowered and "%" not in activity:
            return f"Project Management ({round(pct.pm_pct * 100)}% of Pre-Eng + Engineering)"
        return activity or "Activity"

    def _build_sheet1_costing(self, wb: Workbook, req: AxcendExcelExportRequest):
        ws = wb.create_sheet("Costing")
        ws.sheet_view.showGridLines = True
        border = _thin_border()

        ws.merge_cells("A1:E1")
        ws["A1"] = "Overall Software Design Efforts"
        _style_merged_range(ws, 1, 1, 1, 5, font=_font(bold=True, size=12), align=_align("center"), border=border)

        ws["D3"] = "In Days:"
        _apply(ws["D3"], font=_font(bold=True), align=_align("right"))
        ws["E3"] = "=C13"
        _apply(ws["E3"], font=_font(bold=True), align=_align("center"), border=border)

        # Count unique actual developers in cost_rows to display dynamic number of engineers
        num_engineers = sum(row.count for row in req.cost_rows if not row.is_pm)
        if num_engineers <= 0:
            unique_devs = set()
            for module in req.modules:
                for feature in module.features:
                    if feature.developer:
                        dev_name = feature.developer.split("(")[0].strip()
                        if dev_name.upper() not in {"S1", "S2", "S3", "DEVELOPER", "TESTER"}:
                            unique_devs.add(dev_name)
            num_engineers = len(unique_devs) if unique_devs else 3

        headers_top = ["Man Days", "Man Months", "Man Weeks", f"{num_engineers} Engineers", "Weeks/Eng"]
        for col, header in enumerate(headers_top, start=1):
            _apply(ws.cell(row=5, column=col, value=header), font=_font(bold=True, size=9), align=_align("center"), border=border)

        ws["A6"] = "=ROUND(C13, 0)"
        ws["B6"] = "=ROUND(A6/20, 0)"
        ws["C6"] = "=ROUND(A6/5, 0)"
        ws["D6"] = num_engineers
        ws["E6"] = "=ROUND(C6/D6, 0)"
        for col in range(1, 6):
            _apply(ws.cell(row=6, column=col), font=_font(), align=_align("center"), border=border)

        headers = ["Category", "Man Hrs", "Man Days", f"Rate / Day ({req.currency})", f"Total Dev Cost ({req.currency})"]
        for col, header in enumerate(headers, start=1):
            _apply(ws.cell(row=7, column=col, value=header), font=_font(bold=True), align=_align("center"), border=border)

        s3_rate = s2_rate = s1_rate = 0.0
        for row in req.cost_rows:
            if row.s_level == "S3":
                s3_rate = row.rate_per_day
            elif row.s_level == "S2":
                s2_rate = row.rate_per_day
            elif row.s_level == "S1":
                s1_rate = row.rate_per_day

        # Pull labels from cost_rows to show actual member names (e.g. "Rahul (S2 Dev)")
        cr_map = {r.s_level: r for r in req.cost_rows}
        s3_cr_label = cr_map["S3"].role if "S3" in cr_map else "Sr. Automation Engg. (S3)"
        s2_cr_label = cr_map["S2"].role if "S2" in cr_map else "Automation Engg. (S2)"
        s1_cr_label = cr_map["S1"].role if "S1" in cr_map else "Jr. Automation Engg. (S1)"
        values = [
            (8,  s3_cr_label, "=ROUND('Effort Summary'!B4, 0)", int(round(s3_rate))),
            (9,  s2_cr_label, "=ROUND('Effort Summary'!B5, 0)", int(round(s2_rate))),
            (10, s1_cr_label, "=ROUND('Effort Summary'!B6, 0)", int(round(s1_rate))),
        ]
        for row_num, label, hrs_formula, rate in values:
            ws[f"A{row_num}"] = label
            ws[f"B{row_num}"] = hrs_formula
            ws[f"C{row_num}"] = f"=ROUNDUP(B{row_num}/8, 0)"
            ws[f"D{row_num}"] = rate
            ws[f"E{row_num}"] = f"=ROUND(C{row_num}*D{row_num}, 0)"
            for col in range(1, 6):
                _apply(ws.cell(row=row_num, column=col), font=_font(bold=(col == 5)), align=_align("left" if col == 1 else "center"), border=border)

        ws["A11"] = "TOTAL EFFORTS"
        ws["B11"] = "=SUM(B8:B10)"
        ws["C11"] = "=SUM(C8:C10)"
        ws["E11"] = "=SUM(E8:E10)"
        for col in range(1, 6):
            _apply(ws.cell(row=11, column=col), font=_font(bold=True), fill=_fill(C_SUMMARY_GREY), align=_align("left" if col == 1 else "center"), border=border)

        pm_factor = req.pm_pct / 100.0
        ws["A12"] = f"Project Management ({int(req.pm_pct)}%)"
        ws["B12"] = f"=ROUND(B11*{pm_factor}, 0)"   # PM hours = total hrs × pm%
        ws["C12"] = "=ROUNDUP(B12/8, 0)"             # PM man-days
        ws["D12"] = int(round(s3_rate))               # rate reference (display only)
        # PM cost = devSubtotal × pm% — matches the tool's: pmCost = Math.round(devSubtotal * (pmPct / 100))
        ws["E12"] = f"=ROUND(E11*{pm_factor}, 0)"
        for col in range(1, 6):
            _apply(ws.cell(row=12, column=col), font=_font(), align=_align("left" if col == 1 else "center"), border=border)

        ws["A13"] = "TOTAL EFFORTS"
        ws["B13"] = "=SUM(B11:B12)"
        ws["C13"] = "=SUM(C11:C12)"
        ws["D13"] = "Total Cost"
        ws["E13"] = "=SUM(E11:E12)"
        for col in range(1, 6):
            _apply(ws.cell(row=13, column=col), font=_font(bold=True), fill=_fill(C_GRAND_TOTAL), align=_align("left" if col == 1 else "center"), border=border)

        finance_rows = [
            (15, "Credit period", None),
            (16, f"Finance Cost ({req.finance_cost_pct}%)", f"=ROUND(E13*{req.finance_cost_pct/100}, 0)"),
            (17, f"Forex risk ({req.forex_risk_pct}%)", f"=ROUND(E13*{req.forex_risk_pct/100}, 0)"),
            (18, f"Risk ({req.risk_pct}%)", f"=ROUND(E13*{req.risk_pct/100}, 0)"),
            (19, "SubTotal", "=ROUND(E13+E16+E17+E18, 0)"),
            (20, f"Nego Deduction ({req.nego_deduction_pct}%)", f"=ROUND(E19*{req.nego_deduction_pct/100}, 0)"),
            (21, "FINAL QUOTE", "=ROUND(E19-E20, 0)"),
        ]
        for row_num, label, formula in finance_rows:
            fill = _fill(C_GRAND_TOTAL if label == "FINAL QUOTE" else C_SUMMARY_GREY if label == "SubTotal" else "FFFFFF")
            font = _font(bold=True, size=11 if label == "FINAL QUOTE" else 10)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
            ws.cell(row=row_num, column=1, value=label)
            _style_merged_range(ws, row_num, 1, row_num, 4, font=font, fill=fill, align=_align("right"), border=border)
            _apply(ws.cell(row=row_num, column=5, value=formula), font=font, fill=fill, align=_align("center"), border=border)

        ws.merge_cells("A22:D22")
        ws["A22"] = "Effective Rate/hr:"
        _style_merged_range(ws, 22, 1, 22, 4, font=_font(bold=True), align=_align("right"), border=border)
        _apply(ws.cell(row=22, column=5, value="=ROUND(E21/B13, 0)"), font=_font(bold=True), align=_align("center"), border=border)

        _set_col_widths(ws, {"A": 32, "B": 14, "C": 14, "D": 20, "E": 20})

    def _build_sheet2_effort_summary(self, wb: Workbook, layout: dict[str, object], req: "AxcendExcelExportRequest"):
        ws = wb.create_sheet("Effort Summary")
        ws.sheet_view.showGridLines = True
        border = _thin_border()

        ws.merge_cells("A1:C1")
        ws["A1"] = "Software Engineering"
        _style_merged_range(ws, 1, 1, 1, 3, font=_font(bold=True, size=12), align=_align("center"), border=border)

        for col, header in enumerate(["Software Efforts", "Effort-hrs", "Effort-days"], start=1):
            _apply(ws.cell(row=3, column=col, value=header), font=_font(bold=True), align=_align("center"), border=border)

        # Use hours directly from cost_rows — these come from the frontend display
        # state values and are guaranteed to match what the tool shows on screen.
        # cost_rows already has: S3 = preEng+s3Dev+deployment, S2 = s2Dev+testing, S1 = s1Dev
        cost_row_map = {r.s_level: r for r in req.cost_rows}
        s3_display_hours = int(round(cost_row_map["S3"].hours_per_member)) if "S3" in cost_row_map else 0
        s2_display_hours = int(round(cost_row_map["S2"].hours_per_member)) if "S2" in cost_row_map else 0
        s1_display_hours = int(round(cost_row_map["S1"].hours_per_member)) if "S1" in cost_row_map else 0

        s3_label = cost_row_map["S3"].role if "S3" in cost_row_map else "Sr. Automation Engg. (S3)"
        s2_label = cost_row_map["S2"].role if "S2" in cost_row_map else "Automation Engg. (S2)"
        s1_label = cost_row_map["S1"].role if "S1" in cost_row_map else "Jr. Automation Engg. (S1)"

        summary_rows = [
            (4, s3_label, s3_display_hours),
            (5, s2_label, s2_display_hours),
            (6, s1_label, s1_display_hours),
        ]
        for row_num, label, hours in summary_rows:
            ws[f"A{row_num}"] = label
            ws[f"B{row_num}"] = hours
            ws[f"C{row_num}"] = f"=ROUNDUP(B{row_num}/8, 0)"
            for col in range(1, 4):
                _apply(ws.cell(row=row_num, column=col), font=_font(), align=_align("left" if col == 1 else "center"), border=border)

        ws["A7"] = "Total"
        ws["B7"] = "=SUM(B4:B6)"
        ws["C7"] = "=SUM(C4:C6)"
        for col in range(1, 4):
            _apply(ws.cell(row=7, column=col), font=_font(bold=True), fill=_fill(C_SUMMARY_GREY), align=_align("left" if col == 1 else "center"), border=border)

        _set_col_widths(ws, {"A": 32, "B": 16, "C": 16})

    def _build_sheet3_sw_effort(self, wb: Workbook, req: AxcendExcelExportRequest, layout: dict[str, object]):
        ws = wb.create_sheet("SW Effort")
        ws.sheet_view.showGridLines = True
        border = _thin_border()

        ws.merge_cells("A1:F1")
        ws["A1"] = "AXCEND EFFORT ESTIMATION / FOR FIXED PRICE PROJECTS"
        _style_merged_range(ws, 1, 1, 1, 6, font=_font(bold=True, size=12), align=_align("center"), border=border)

        # Legend row A2:F2 removed as requested

        headers = ["Activity Planned", "Resources (Location)", "Resource Level", "Years exp", "Milestone/Phase", "Input Hours"]
        for col, header in enumerate(headers, start=1):
            _apply(ws.cell(row=4, column=col, value=header), font=_font(bold=True), align=_align("center"), border=border)

        # Build a level → display name map from cost_rows so Resource Level shows
        # the actual developer name (e.g. "Anjana R (S1 Dev)") instead of bare codes.
        level_name_map: dict[str, str] = {}
        for cr in (req.cost_rows or []):
            if cr.s_level and cr.role:
                level_name_map[cr.s_level] = cr.role

        def _resource_display(item) -> str:
            """Return 'Name (Sn)' if we have a roster name, else keep the raw level."""
            level = item.resource_level or ""
            name = level_name_map.get(level, "")
            if name:
                return f"{name}"
            return level

        def write_section_header(row_num: int, label: str):
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
            ws.cell(row=row_num, column=1, value=label)
            _style_merged_range(ws, row_num, 1, row_num, 6, font=_font(bold=True), align=_align("left"), border=border)

        write_section_header(5, "PRE-ENGINEERING")
        for row_num, item in layout["pre_rows"]:
            values = [self._activity_label(item, req), item.location, _resource_display(item), item.experience_years, "", int(round(item.input_hours))]
            for col, value in enumerate(values, start=1):
                _apply(ws.cell(row=row_num, column=col, value=value), font=_font(), align=_align("left" if col == 1 else "center"), border=border)

        pre_rows = [row_num for row_num, _ in layout["pre_rows"]]
        ws.merge_cells(start_row=layout["pre_total_row"], start_column=1, end_row=layout["pre_total_row"], end_column=5)
        ws.cell(row=layout["pre_total_row"], column=1, value="Pre-Engineering Total")
        _style_merged_range(ws, layout["pre_total_row"], 1, layout["pre_total_row"], 5, font=_font(bold=True), fill=_fill(C_SUMMARY_GREY), align=_align("right"), border=border)
        pre_formula = f"=ROUND(SUM(F{pre_rows[0]}:F{pre_rows[-1]}), 0)" if pre_rows else "=0"
        _apply(ws.cell(row=layout["pre_total_row"], column=6, value=pre_formula), font=_font(bold=True), fill=_fill(C_SUMMARY_GREY), align=_align("center"), border=border)

        write_section_header(layout["pre_total_row"] + 2, "DEVELOPMENT")

        internal_pct = req.effort_percentages.internal_testing_pct
        client_pct = req.effort_percentages.client_testing_pct
        deploy_pct = req.effort_percentages.deployment_pct

        for idx, (row_num, item) in enumerate(layout["eng_rows"]):
            # Use the hours from the item directly — these come from cost_rows which
            # are derived from the exact display state values (preEngHours, s1Hours,
            # s2Hours, s3DevHours) so they always match the on-screen tool output.
            hours_value = int(round(item.input_hours)) if item.input_hours else 0

            values = [self._activity_label(item, req), item.location, _resource_display(item), item.experience_years, "", hours_value]
            for col, value in enumerate(values, start=1):
                _apply(ws.cell(row=row_num, column=col, value=value), font=_font(), align=_align("left" if col == 1 else "center"), border=border)

        eng_rows_nums = [row_num for row_num, _ in layout["eng_rows"]]
        ws.merge_cells(start_row=layout["dev_total_row"], start_column=1, end_row=layout["dev_total_row"], end_column=5)
        ws.cell(row=layout["dev_total_row"], column=1, value="Development Total")
        _style_merged_range(ws, layout["dev_total_row"], 1, layout["dev_total_row"], 5, font=_font(bold=True), fill=_fill(C_SUMMARY_GREY), align=_align("right"), border=border)
        eng_formula = f"=ROUND(SUM(F{eng_rows_nums[0]}:F{eng_rows_nums[-1]}), 0)" if eng_rows_nums else "=0"
        _apply(ws.cell(row=layout["dev_total_row"], column=6, value=eng_formula), font=_font(bold=True), fill=_fill(C_SUMMARY_GREY), align=_align("center"), border=border)

        # Project Management is NOT included in SW Effort sheet (only in Costing sheet)

        ws.merge_cells(start_row=layout["grand_total_row"], start_column=1, end_row=layout["grand_total_row"], end_column=5)
        ws.cell(row=layout["grand_total_row"], column=1, value="GRAND TOTAL")
        _style_merged_range(ws, layout["grand_total_row"], 1, layout["grand_total_row"], 5, font=_font(bold=True, size=11), fill=_fill(C_GRAND_TOTAL), align=_align("right"), border=border)
        _apply(ws.cell(row=layout["grand_total_row"], column=6, value=f"=ROUND(F{layout['pre_total_row']}+F{layout['dev_total_row']}, 0)"), font=_font(bold=True, size=11), fill=_fill(C_GRAND_TOTAL), align=_align("center"), border=border)

        _set_col_widths(ws, {"A": 42, "B": 22, "C": 30, "D": 14, "E": 18, "F": 16})

    def _build_sheet4_software_efforts(self, wb: Workbook, req: AxcendExcelExportRequest):
        ws = wb.create_sheet("Software Efforts")
        ws.sheet_view.showGridLines = True
        border = _thin_border()

        ws.merge_cells("A1:E1")
        ws["A1"] = f"Software Efforts - {req.project_name}"
        _style_merged_range(ws, 1, 1, 1, 5, font=_font(bold=True, size=12), align=_align("center"), border=border)

        for col, header in enumerate(["SL", "Module", "Feature", "Description", "Estimated Hours"], start=1):
            _apply(ws.cell(row=2, column=col, value=header), font=_font(bold=True, size=11), align=_align("center"), border=border)

        row = 3
        for module in req.modules:
            module_start = row
            for feature in module.features:
                # Use estimated_hours directly — the frontend already applied the correct
                # experience multiplier per roster member. Re-deriving from base_hours
                # breaks for real person names (not "S1"/"S2"/"S3" literals).
                est_hrs = int(round(feature.estimated_hours)) if feature.estimated_hours > 0 else 0

                values = [feature.sl, feature.module, feature.feature, feature.description or "", est_hrs]
                for col, value in enumerate(values, start=1):
                    _apply(ws.cell(row=row, column=col, value=value), font=_font(), align=_align("center" if col in (1, 5) else "left", "top"), border=border)
                row += 1
            if row - 1 > module_start:
                ws.merge_cells(start_row=module_start, start_column=2, end_row=row - 1, end_column=2)
                _style_merged_range(ws, module_start, 2, row - 1, 2, font=_font(), align=_align("left", "top"), border=border)

        dd_total_row = row
        internal_pct = req.effort_percentages.internal_testing_pct
        client_pct = req.effort_percentages.client_testing_pct
        deploy_pct = req.effort_percentages.deployment_pct

        # Extract actual testing/deployment hours from the engineering rows so
        # Sheet 4 always shows the same values as the on-screen tool, even when
        # the D&D feature list is empty (avoids formula-chains producing 0).
        def _eng_actual_hrs(keyword: str) -> int:
            return int(round(sum(
                r.input_hours for r in req.engineering
                if keyword in (r.activity or "").lower() and r.input_hours
            )))

        internal_testing_hrs = _eng_actual_hrs("internal testing")
        client_testing_hrs   = _eng_actual_hrs("client testing")
        deployment_hrs       = _eng_actual_hrs("deployment")

        summary_rows = [
            (f"Design and Development", f"=ROUND(SUM(E3:E{dd_total_row - 1}), 0)", True),
            (f"Internal Testing ({round(internal_pct * 100)}% of D&D)", internal_testing_hrs, False),
            (f"Client Testing ({round(client_pct * 100)}% of D&D)", client_testing_hrs, False),
            (f"Deployment ({round(deploy_pct * 100)}% of D&D)", deployment_hrs, False),
        ]
        for label, formula, is_summary in summary_rows:
            fill = _fill(C_SUMMARY_GREY) if is_summary else None
            font = _font(bold=True) if is_summary else _font()
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value=label)
            _style_merged_range(ws, row, 1, row, 4, font=font, fill=fill, align=_align("right"), border=border)
            _apply(ws.cell(row=row, column=5, value=formula), font=font, fill=fill, align=_align("center"), border=border)
            row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value="GRAND TOTAL")
        _style_merged_range(ws, row, 1, row, 4, font=_font(bold=True, size=11), fill=_fill(C_GRAND_TOTAL), align=_align("right"), border=border)
        _apply(ws.cell(row=row, column=5, value=f"=ROUND(SUM(E{dd_total_row}:E{row - 1}), 0)"), font=_font(bold=True, size=11), fill=_fill(C_GRAND_TOTAL), align=_align("center"), border=border)

        _set_col_widths(ws, {"A": 6, "B": 24, "C": 32, "D": 60, "E": 18})
